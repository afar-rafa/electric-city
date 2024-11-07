import csv
import logging
import os
from typing import Dict, List, Union

import openpyxl

from classes.edificio import Edificio
from helpers.constants import OUTPUT_FORMAT, script_dir

CSV_QUOTECHAR = '"'
OUTPUT_FOLDER = f"{script_dir}/outputs"

logger = logging.getLogger(__name__)


# Clase base para manejar archivos
class DBFileHandler:
    def __init__(self):
        # Almacena las filas en memoria para cada archivo
        self.file_buffers: Dict[str, List[List[Union[str, int, float]]]] = {}

    def crear_archivo(self, nombre: str, headers: List[str]):
        raise NotImplementedError

    def agregar_fila_en_memoria(self, nombre: str, fila: List[Union[str, int, float]]):
        if nombre not in self.file_buffers:
            self.file_buffers[nombre] = []
        self.file_buffers[nombre].append(fila)

    def exportar_archivos(self):
        raise NotImplementedError

    def leer(self, nombre: str):
        raise NotImplementedError

    def leer_headers(self, nombre: str):
        raise NotImplementedError


# Para archivos CSV
class CSVFileHandler(DBFileHandler):
    def __init__(self, format):
        super().__init__()
        if format == ".tsv":
            self.CSV_DELIMITER = "\t"
        else:
            self.CSV_DELIMITER = ","

    def crear_archivo(self, nombre: str, headers: List[str]):
        with open(nombre, "w") as csv_file:
            csv_writer = csv.writer(
                csv_file,
                delimiter=self.CSV_DELIMITER,
                quotechar=CSV_QUOTECHAR,
            )
            csv_writer.writerow(headers)

    def exportar_archivos(self):
        for nombre, filas in self.file_buffers.items():
            logger.warning(f"DB - guardando '{nombre}'")
            with open(nombre, "a") as csv_file:
                csv_writer = csv.writer(
                    csv_file,
                    delimiter=self.CSV_DELIMITER,
                    quotechar=CSV_QUOTECHAR,
                )
                csv_writer.writerows(filas)

    def leer(
        self,
        nombre: str,
    ):
        with open(nombre, newline="") as csv_file:
            spamreader = csv.DictReader(
                csv_file,
                delimiter=self.CSV_DELIMITER,
                quotechar=CSV_QUOTECHAR,
            )
            for row in spamreader:
                yield row

    def leer_headers(self, nombre: str):
        with open(nombre, newline="") as csv_file:
            spamreader = csv.reader(
                csv_file,
                delimiter=self.CSV_DELIMITER,
                quotechar=CSV_QUOTECHAR,
            )
            for row in spamreader:
                return row


# Paras archivos Excel (.xlsx)
class ExcelFileHandler(DBFileHandler):
    def crear_archivo(self, nombre: str, headers: List[str]):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(nombre)

    def exportar_archivos(self):
        for nombre, filas in self.file_buffers.items():
            logger.warning(f"DB - guardando '{nombre}'")
            wb = openpyxl.load_workbook(nombre)
            ws = wb.active
            for fila in filas:
                ws.append(fila)
            wb.save(nombre)

    def leer(self, nombre: str):
        wb = openpyxl.load_workbook(nombre)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            yield dict(
                # Assuming first row as header
                zip([cell.value for cell in ws[1]], row)
            )

    def leer_headers(self, nombre: str):
        wb = openpyxl.load_workbook(nombre)
        ws = wb.active
        # First row as headers
        return [cell.value for cell in ws[1]]


# Clase principal que selecciona el lector de archivos adecuado
class DB:
    handler = None

    def __init__(self, extension: str = None):
        if extension:
            self.cambiar_handler(extension)

    def _get_handler(self, file_name: str) -> DBFileHandler:
        if not self.handler:
            extension = os.path.splitext(file_name)[1].lower()
            self.cambiar_handler(extension)
        return self.handler

    def cambiar_handler(self, extension: str):
        logger.info("Simulación - Usando archivos %s", extension)
        if extension == ".csv" or extension == ".tsv":
            self.handler = CSVFileHandler(extension)
        elif extension == ".xlsx":
            self.handler = ExcelFileHandler()
        else:
            raise ValueError(f"Unsupported file extension: {extension}")

    def crear_archivo(self, nombre: str, headers: List[str]):
        handler = self._get_handler(nombre)
        handler.crear_archivo(nombre, headers)

    def agregar_fila_en_memoria(self, nombre: str, fila: List[Union[str, int, float]]):
        handler = self._get_handler(nombre)
        handler.agregar_fila_en_memoria(nombre, fila)

    def exportar_archivos(self):
        if self.handler:
            self.handler.exportar_archivos()

    def leer(self, nombre: str):
        handler = self._get_handler(nombre)
        return handler.leer(nombre)

    def leer_headers(self, nombre: str):
        handler = self._get_handler(nombre)
        return handler.leer_headers(nombre)

    def crear_archivo_de_edificios(self, edificios: List["Edificio"]):  # type: ignore
        for e in edificios:
            self.crear_archivo(
                nombre=f"{OUTPUT_FOLDER}/{e}.{OUTPUT_FORMAT}",
                headers=["Tiempo", "Potencia Disponible", "Gasto de Cargadores"]
                + [f"{v}" for v in e.vehículos],
            )
            if e.tipo_edificio == Edificio.TIPO_INT:
                self.crear_archivo(
                    nombre=f"{OUTPUT_FOLDER}/Prioridades {e}.{OUTPUT_FORMAT}",
                    headers=["Tiempo"] + [f"{v}" for v in e.vehículos],
                )

    def guardar_estado_de_edificio(self, tiempo: str, e: Edificio):
        fila = [tiempo, e.potencia_disponible, e.potencia_usada_por_autos] + e.bateria_de_vehículos

        logger.info("Simulación: %s", fila)
        self.agregar_fila_en_memoria(f"{OUTPUT_FOLDER}/{e}.{OUTPUT_FORMAT}", fila)

        if e.tipo_edificio == Edificio.TIPO_INT:
            fila_prioridades = [tiempo] + e.prioridad_de_vehículos
            self.agregar_fila_en_memoria(
                f"{OUTPUT_FOLDER}/Prioridades {e}.{OUTPUT_FORMAT}",
                fila_prioridades,
            )
